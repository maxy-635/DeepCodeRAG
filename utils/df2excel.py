import os
import openpyxl


class DataFrame2Excel:
    """
    将DataFrame数据保存到excel文件中
    :param data: DataFrame数据
    :param excel_path: excel文件路径
    :return: None
    """

    def __init__(self, data, save_excel_path):
        self.data = data
        self.save_excel_path = save_excel_path

    def df2excel(self, sheet_name):
        # 打开已存在的 Excel 文件，如果不存在则创建一个新的 Workbook 对象
        if os.path.exists(self.save_excel_path):
            wb = openpyxl.load_workbook(self.save_excel_path)
        else:
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)  # 删除默认的第一个多余的sheet

        # 创建一个新的 sheet，并指定 sheet 名称为 prompting
        sheet = wb.create_sheet(title=sheet_name)
        # 写入列名
        for col_idx, col_name in enumerate(self.data.columns, start=1):
            sheet.cell(row=1, column=col_idx, value=col_name)
        # 写入数据行
        for r_idx, row in enumerate(self.data.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        if not os.path.exists(os.path.dirname(self.save_excel_path)):
            os.makedirs(os.path.dirname(self.save_excel_path))
        wb.save(self.save_excel_path)
